from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from portal.models import InventoryItem


class Command(BaseCommand):
    help = (
        "Импортирует справочник товаров из Excel (inventory.xlsx). "
        "Ожидаемые колонки: SKU, Название, Категория, Ед., Упаковка, Маркировка упаковки, Цена, Остаток, Локация, Примечание"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            dest="file_path",
            default="data/inventory.xlsx",
            help="Путь к Excel-файлу со справочником товаров",
        )

    def _normalize_unit(self, raw):
        if not raw:
            return None
        raw = str(raw).strip().lower()
        mapping = {
            "м": InventoryItem.Unit.METER,
            "meter": InventoryItem.Unit.METER,
            "м2": InventoryItem.Unit.SQUARE_METER,
            "кв.м": InventoryItem.Unit.SQUARE_METER,
            "sqm": InventoryItem.Unit.SQUARE_METER,
            "лист": InventoryItem.Unit.SHEET,
            "листы": InventoryItem.Unit.SHEET,
            "sheet": InventoryItem.Unit.SHEET,
            "шт": InventoryItem.Unit.PIECE,
            "pcs": InventoryItem.Unit.PIECE,
        }
        return mapping.get(raw, InventoryItem.Unit.PIECE)

    @transaction.atomic
    def handle(self, *args, **options):
        path = options["file_path"]
        try:
            wb = load_workbook(path)
        except FileNotFoundError as exc:
            raise CommandError(f"Файл не найден: {path}") from exc

        ws = wb.active
        if ws.max_row < 2:
            self.stdout.write(self.style.WARNING("В файле нет данных"))
            return

        created = 0
        updated = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            sku, name, category, unit_raw, package_size, package_label, price, qty, location, notes = (
                (row[i] if i < len(row) else None) for i in range(10)
            )

            if not name:
                self.stdout.write(self.style.WARNING(f"Строка {idx}: пропущена (нет названия)"))
                continue

            defaults = {
                "name": str(name).strip(),
                "category": category or "",
                "base_unit": self._normalize_unit(unit_raw) or InventoryItem.Unit.PIECE,
                "package_size": package_size or None,
                "package_unit_label": package_label or "",
                "default_unit_price": price or 0,
                "quantity_on_hand": qty or 0,
                "location": location or "",
                "notes": notes or "",
            }

            sku = str(sku).strip() if sku else None
            if not sku:
                self.stdout.write(self.style.WARNING(f"Строка {idx}: пропущена (нет SKU)"))
                continue

            item, created_flag = InventoryItem.objects.update_or_create(
                sku=sku, defaults=defaults
            )
            created += int(created_flag)
            updated += int(not created_flag)

        self.stdout.write(self.style.SUCCESS(f"Создано: {created}, обновлено: {updated}"))
