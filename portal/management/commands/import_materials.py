from datetime import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from portal.models import InventoryItem, InventoryUsage, Order


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str) and value:
        try:
            return datetime.strptime(value, "%d.%m.%Y").date()
        except ValueError:
            return datetime.fromisoformat(value).date()
    return None


def _normalize_unit(unit_value: str):
    if not unit_value:
        return None
    unit_value = str(unit_value).strip().lower()
    mapping = {
        "м": InventoryItem.Unit.METER,
        "метр": InventoryItem.Unit.METER,
        "метры": InventoryItem.Unit.METER,
        "шт": InventoryItem.Unit.PIECE,
        "штука": InventoryItem.Unit.PIECE,
        "штуки": InventoryItem.Unit.PIECE,
        "лист": InventoryItem.Unit.SHEET,
        "листы": InventoryItem.Unit.SHEET,
        "рулон": InventoryItem.Unit.METER,
        "кв.м": InventoryItem.Unit.SQUARE_METER,
        "кв метр": InventoryItem.Unit.SQUARE_METER,
    }
    return mapping.get(unit_value)


def _get_or_create_item(name: str, unit_label: str | None):
    existing = InventoryItem.objects.filter(name__iexact=name).first()
    if existing:
        return existing, False

    base_sku = slugify(name) or "item"
    sku = base_sku
    counter = 1
    while InventoryItem.objects.filter(sku=sku).exists():
        counter += 1
        sku = f"{base_sku}-{counter}"

    base_unit = _normalize_unit(unit_label) or InventoryItem.Unit.PIECE
    created = InventoryItem.objects.create(name=name, sku=sku, base_unit=base_unit)
    return created, True


class Command(BaseCommand):
    help = (
        "Импортирует расход материалов из Excel (materials.xlsx) в таблицу InventoryUsage. "
        "Формат колонок: Дата, Товар, Количество, Ед. изм., Проект"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            dest="file_path",
            default="data/materials.xlsx",
            help="Путь к файлу Excel с расходами материалов",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = options["file_path"]
        try:
            wb = load_workbook(path)
        except FileNotFoundError as exc:
            raise CommandError(f"Файл не найден: {path}") from exc

        ws = wb.active
        if ws.max_row < 2:
            self.stdout.write(self.style.WARNING("В файле нет данных"))
            return

        created_usages = 0
        created_items = 0
        skipped_rows = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            usage_date_raw, product_name, quantity_raw, unit_raw, project_raw = row[:5]

            if not product_name or not quantity_raw:
                skipped_rows += 1
                self.stdout.write(self.style.WARNING(f"Строка {idx}: пропущена (нет товара или количества)"))
                continue

            usage_date = _parse_date(usage_date_raw) or datetime.today().date()
            try:
                quantity = float(quantity_raw)
            except (TypeError, ValueError):
                skipped_rows += 1
                self.stdout.write(self.style.WARNING(f"Строка {idx}: некорректное количество {quantity_raw}"))
                continue

            item, is_new = _get_or_create_item(str(product_name).strip(), unit_raw)
            if is_new:
                created_items += 1

            project = None
            if project_raw:
                project = Order.objects.filter(title__iexact=str(project_raw).strip()).first()

            InventoryUsage.objects.create(
                usage_date=usage_date,
                item=item,
                quantity=quantity,
                project=project,
                comment="Импорт из materials.xlsx",
            )
            created_usages += 1

        self.stdout.write(self.style.SUCCESS(f"Создано списаний: {created_usages}"))
        if created_items:
            self.stdout.write(self.style.SUCCESS(f"Создано новых товаров: {created_items}"))
        if skipped_rows:
            self.stdout.write(self.style.WARNING(f"Пропущено строк: {skipped_rows}"))
